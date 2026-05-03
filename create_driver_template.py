"""Run once to generate driver_template.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Drivers"

HEADERS = [
    "drivercode", "name", "phone", "home_city", "region",
    "vehicle_type", "works_shabbat", "works_nights",
    "works_long_distance", "languages", "active", "notes"
]

INSTRUCTIONS = [
    "Unique ID (matches supplier system)",
    "Full name",
    "WhatsApp number (e.g. 0521234567)",
    "City driver is based in",
    "Region (e.g. Center, North, South)",
    "sedan / executive_minivan / minivan / minibus_15 / minibus_18",
    "TRUE or FALSE",
    "TRUE or FALSE",
    "TRUE or FALSE",
    "Comma-separated (e.g. Hebrew,English)",
    "TRUE or FALSE",
    "Free text"
]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
instruction_fill = PatternFill("solid", fgColor="D6E4F0")
instruction_font = Font(italic=True, color="555555", size=9)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, (header, instruction) in enumerate(zip(HEADERS, INSTRUCTIONS), start=1):
    hc = ws.cell(row=1, column=col, value=header)
    hc.font = header_font
    hc.fill = header_fill
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hc.border = border

    ic = ws.cell(row=2, column=col, value=instruction)
    ic.font = instruction_font
    ic.fill = instruction_fill
    ic.alignment = Alignment(wrap_text=True)
    ic.border = border

# Sample rows
samples = [
    ["DR001", "Moshe Cohen", "0521234567", "Tel Aviv", "Center",
     "sedan", "FALSE", "FALSE", "FALSE", "Hebrew,English", "TRUE", ""],
    ["DR002", "David Levy", "0527654321", "Haifa", "North",
     "executive_minivan", "TRUE", "TRUE", "TRUE", "Hebrew", "TRUE", "Prefers airport jobs"],
    ["DR003", "Yossi Mizrahi", "0531112222", "Beer Sheva", "South",
     "minivan", "FALSE", "TRUE", "FALSE", "Hebrew,Russian", "TRUE", ""],
    ["DR004", "Avi Peretz", "0543334444", "Jerusalem", "Center",
     "minibus_15", "TRUE", "FALSE", "TRUE", "Hebrew,Arabic", "TRUE", "Group tours specialist"],
]

for row_idx, row_data in enumerate(samples, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="center")

# Vehicle type dropdown validation
vt_dv = DataValidation(
    type="list",
    formula1='"sedan,executive_minivan,minivan,minibus_15,minibus_18"',
    allow_blank=False,
    showDropDown=False,
)
ws.add_data_validation(vt_dv)
vt_dv.sqref = f"F3:F1000"

# Bool dropdowns
bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
ws.add_data_validation(bool_dv)
bool_dv.sqref = "G3:G1000 H3:H1000 I3:I1000 K3:K1000"

# Column widths
widths = [15, 20, 16, 16, 12, 20, 15, 14, 18, 22, 8, 30]
for col, width in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 40
ws.freeze_panes = "A3"

wb.save("driver_template.xlsx")
print("driver_template.xlsx created.")
